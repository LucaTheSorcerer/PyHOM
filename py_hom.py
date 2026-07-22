import numpy as np
import argparse

# 1. Define the Lambda values from the table for clamped-clamped beams
lambdas = {
    1: 4.730041,
    2: 7.853205,
    3: 10.995608,
    4: 14.137165,
    5: 17.278760,
    6: 20.420352,
}


def calculate_squared_mode_shape(lambda_val, z_array, length=None):
    """
    Calculates and normalizes the squared mode shape.
    If length is provided, normalizes z_array by length to map to [0, 1].
    """
    # Use normalized coordinate if length is specified, otherwise use z_array directly
    # (Using normalized coordinate is the physically correct way for generic lengths)
    zeta = z_array / length if length is not None else z_array

    numerator = np.cos(lambda_val) - np.cosh(lambda_val)
    denominator = np.sin(lambda_val) - np.sinh(lambda_val)
    sigma = numerator / denominator

    phi = (np.cosh(lambda_val * zeta) - np.cos(lambda_val * zeta)) - \
          sigma * (np.sinh(lambda_val * zeta) - np.sin(lambda_val * zeta))

    phi_normalized = phi / np.max(np.abs(phi))
    return phi_normalized ** 2


def integrate_region(Z_matrix, x_coords, y_coords, x_bounds, y_bounds):
    """Performs a 2D numerical double integral over a specific region using Trapezoidal rule."""
    tol = 1e-5
    mask_x = (x_coords >= x_bounds[0] - tol) & (x_coords <= x_bounds[1] + tol)
    mask_y = (y_coords >= y_bounds[0] - tol) & (y_coords <= y_bounds[1] + tol)

    if not np.any(mask_x) or not np.any(mask_y):
        return 0.0

    Z_sub = Z_matrix[mask_x][:, mask_y]
    x_sub = x_coords[mask_x]
    y_sub = y_coords[mask_y]

    if len(x_sub) < 2 or len(y_sub) < 2:
        return 0.0

    integral_y = np.trapz(Z_sub, x=y_sub, axis=1)
    integral_total = np.trapz(integral_y, x=x_sub, axis=0)

    return integral_total


def compute_freq_integral(mode_i, mode_j, X1, X2, A1, B1, A2, B2, f_ref, mass_ratio):
    x_coords = np.linspace(0, X1, 101)
    y_coords = np.linspace(0, X2, 101)

    phi_sq_x = calculate_squared_mode_shape(lambdas[mode_i], x_coords, X1)
    phi_sq_y = calculate_squared_mode_shape(lambdas[mode_j], y_coords, X2)
    Z_matrix = np.outer(phi_sq_x, phi_sq_y)

    phi_total = integrate_region(Z_matrix, x_coords, y_coords, [0, X1], [0, X2])
    phi_defect = integrate_region(Z_matrix, x_coords, y_coords, [A1, B1], [A2, B2])

    denom = (phi_total - phi_defect) + phi_defect * mass_ratio
    freq_ratio = np.sqrt(phi_total / denom) if denom > 0 else 0
    return f_ref * freq_ratio


def process_batch_excel(input_excel='Correct_data.xlsx', output_excel='Initial_results_Formatted.xlsx'):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    print(f"Reading from {input_excel}...")
    wb_input = openpyxl.load_workbook(input_excel, data_only=True)
    ws_input = wb_input['Sheet1']

    # Parse mode combinations
    modes = []
    for c in range(4, 24):
        mode_str = str(ws_input.cell(row=2, column=c).value).strip()
        parts = mode_str.split('-')
        modes.append((int(parts[0].strip()), int(parts[1].strip())))

    # Parse healthy frequencies
    healthy_freqs = [float(ws_input.cell(row=3, column=c).value) for c in range(4, 24)]

    # Parse test cases
    test_cases = []
    for r in range(5, 9):
        name = ws_input.cell(row=r, column=1).value
        x_range = str(ws_input.cell(row=r, column=2).value).split('-')
        y_range = str(ws_input.cell(row=r, column=3).value).split('-')

        A1, B1 = float(x_range[0]), float(x_range[1])
        A2, B2 = float(y_range[0]), float(y_range[1])
        start_x_mm, start_y_mm = int(A1 * 1000), int(A2 * 1000)

        fem_freqs = [float(ws_input.cell(row=r, column=c).value) if ws_input.cell(row=r, column=c).value else None for c
                     in range(4, 24)]

        test_cases.append({
            'name': name, 'start_x_mm': start_x_mm, 'start_y_mm': start_y_mm,
            'A1': A1, 'B1': B1, 'A2': A2, 'B2': B2, 'fem_freqs': fem_freqs
        })

    # Run computations
    results = []
    X1, X2 = 3.0, 2.0
    mass_ratio = 15000.0 / 7850.0

    print(f"Computing frequencies for {len(test_cases)} test cases x 20 modes...")
    for tc in test_cases:
        for mode_idx in range(20):
            mode_i, mode_j = modes[mode_idx]
            f_ref = healthy_freqs[mode_idx]
            fem_freq = tc['fem_freqs'][mode_idx]

            python_freq = compute_freq_integral(mode_i, mode_j, X1, X2, tc['A1'], tc['B1'], tc['A2'], tc['B2'], f_ref,
                                                mass_ratio)
            abs_error = abs(fem_freq - python_freq) if fem_freq is not None else None

            results.append({
                'healthy_plate': 'L1 = 3 metri' if mode_idx == 0 else ('L2 = 2 metri' if mode_idx == 1 else ''),
                'defect_plate': 'L1 = 2 metri' if mode_idx == 0 else ('L2 = 2 metri' if mode_idx == 1 else ''),
                'test_case': tc['name'],
                'start_x_mm': tc['start_x_mm'], 'start_y_mm': tc['start_y_mm'],
                'A1': tc['A1'], 'B1': tc['B1'], 'A2': tc['A2'], 'B2': tc['B2'],
                'mode': f"{mode_i}-{mode_j}", 'f_healthy': f_ref, 'f_fem': fem_freq,
                'f_python': python_freq, 'abs_error': abs_error
            })

    # Write output
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = 'Validation Results'

    headers = ['Healthy Plate Dimensions', 'Defect Plate Dimensions', 'Test Case', 'Defect Start X (mm)',
               'Defect Start Y (mm)',
               'A1 (m)', 'B1 (m)', 'A2 (m)', 'B2 (m)', 'Mode (i-j)', 'Freq Healthy (Hz)', 'Freq Defect FEM (Hz)',
               'Python Integral Frequency with added mass(Hz)', 'Absolute Error (Hz)']

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font;
        cell.fill = header_fill;
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=r['healthy_plate']).border = thin_border
        ws.cell(row=i, column=2, value=r['defect_plate']).border = thin_border
        ws.cell(row=i, column=3, value=r['test_case']).border = thin_border
        ws.cell(row=i, column=4, value=r['start_x_mm']).border = thin_border
        ws.cell(row=i, column=5, value=r['start_y_mm']).border = thin_border
        ws.cell(row=i, column=6, value=r['A1']).border = thin_border
        ws.cell(row=i, column=7, value=r['B1']).border = thin_border
        ws.cell(row=i, column=8, value=r['A2']).border = thin_border
        ws.cell(row=i, column=9, value=r['B2']).border = thin_border
        ws.cell(row=i, column=10, value=r['mode']).border = thin_border
        ws.cell(row=i, column=11, value=r['f_healthy']).border = thin_border
        ws.cell(row=i, column=12, value=r['f_fem']).border = thin_border
        ws.cell(row=i, column=13, value=round(r['f_python'], 6)).border = thin_border
        ws.cell(row=i, column=14,
                value=round(r['abs_error'], 6) if r['abs_error'] is not None else '').border = thin_border

    col_widths = [22, 22, 10, 18, 18, 8, 8, 8, 8, 12, 16, 18, 40, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    wb_out.save(output_excel)
    print(f"Successfully saved results to {output_excel}")


def main():
    parser = argparse.ArgumentParser(description="2D Plate Vibration Analysis")
    parser.add_argument('--batch', type=str, help="Path to input Excel file for batch processing")
    parser.add_argument('--output', type=str, default='Initial_results_Formatted.xlsx',
                        help="Path to output Excel file")
    args = parser.parse_args()

    if args.batch:
        process_batch_excel(args.batch, args.output)
        return

    print("=== 2D Plate Vibration Analysis CLI ===")

    # We will use simple inputs to make it easy to interact with
    try:
        print("\n--- Plate Dimensions ---")
        X1 = float(input("Enter plate length X1 (e.g., 2): "))
        X2 = float(input("Enter plate length X2 (e.g., 3): "))

        print("\n--- Added Density Plate (Defect) Boundaries ---")
        A1 = float(input(f"Enter A1 (start X, 0 to {X1}): "))
        B1 = float(input(f"Enter B1 (end X, A1 to {X1}): "))
        A2 = float(input(f"Enter A2 (start Y, 0 to {X2}): "))
        B2 = float(input(f"Enter B2 (end Y, A2 to {X2}): "))

        print("\n--- Vibration Modes ---")
        mode_i = int(input("Enter mode i for X axis (1-10): "))
        mode_j = int(input("Enter mode j for Y axis (1-10): "))

        print("\n--- Physical Parameters ---")
        f_ref = float(input("Enter reference frequency without mass (e.g., 24): "))
        mass_ratio = float(
            input("Enter mass ratio in the patch region (e.g., > 1 for heavier plate addition, default 1.1): "))

    except ValueError:
        print("Invalid input. Please enter numbers.")
        return
    except KeyboardInterrupt:
        print("\nOperation cancelled.")
        return

    # Check bounds
    A1, B1 = sorted([A1, B1])
    A2, B2 = sorted([A2, B2])

    # Generate 101 points grid for both axes
    x_coords = np.linspace(0, X1, 101)
    y_coords = np.linspace(0, X2, 101)

    # Calculate mode shapes
    phi_sq_x = calculate_squared_mode_shape(lambdas[mode_i], x_coords, X1)
    phi_sq_y = calculate_squared_mode_shape(lambdas[mode_j], y_coords, X2)

    Z_matrix = np.outer(phi_sq_x, phi_sq_y)

    # Calculate participations via 2D integration
    phi_total_integral = integrate_region(Z_matrix, x_coords, y_coords, [0, X1], [0, X2])
    phi_defect_integral = integrate_region(Z_matrix, x_coords, y_coords, [A1, B1], [A2, B2])

    # Calculate participations via Discrete Sum

    denom_integral = (phi_total_integral - phi_defect_integral) + phi_defect_integral * mass_ratio
    freq_ratio_int = np.sqrt(phi_total_integral / denom_integral) if denom_integral > 0 else 0
    f_with_mass_int = f_ref * freq_ratio_int


    print("\n==============================================")
    print("          METHOD: PHYSICS INTEGRAL          ")
    print("==============================================")
    print("Mass/Integral Statistics:")
    plate_area = X1 * X2
    integral_avg = phi_total_integral / plate_area if plate_area > 0 else 0
    print(f"  Total Plate Integral (Intact)\t{phi_total_integral:.6f}".replace('.', ','))
    print(f"  Defect Region Integral\t{phi_defect_integral:.6f}".replace('.', ','))
    print(f"  Physical Area (m^2)\t\t{plate_area:.6f}".replace('.', ','))
    print(f"  Average per m^2\t\t{integral_avg:.9f}".replace('.', ','))
    print()
    print("Frequency Results:")
    print(f"  10\t\t{10 * freq_ratio_int:.9f}\t{freq_ratio_int:.9f}".replace('.', ','))
    print("  \t\tFrequency changes with\t")
    print()
    print(f"  Frequency with mass\t{f_with_mass_int:.6f} Hz".replace('.', ','))
    print(f"  Frequency without mass\t{f_ref:.6f} Hz".replace('.', ','))
    print("==============================================\n")


if __name__ == '__main__':
    main()